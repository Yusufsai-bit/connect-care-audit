"""
Connect Care Services — NDIS Compliance Dashboard
Run: streamlit run connecteam_dashboard.py
"""

import os, io, calendar, datetime, json, uuid
import pandas as pd
import streamlit as st

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    _OPENPYXL_OK = True
except ImportError:
    _OPENPYXL_OK = False

# ── Secrets → env vars ────────────────────────────────────────────────────────
for _k, _e in [
    ("CONNECTEAM_API_KEY",    "CONNECTEAM_API_KEY"),
    ("ANTHROPIC_API_KEY",     "ANTHROPIC_API_KEY"),
    ("CONNECTEAM_SENDER_ID",  "CONNECTEAM_SENDER_ID"),
]:
    _v = st.secrets.get(_k, "")
    if _v: os.environ[_e] = _v

CT_CHAT_READY = bool(st.secrets.get("CONNECTEAM_SENDER_ID", ""))

from connecteam_audit import (
    run_audit, fetch_all_users,
    fetch_worker_credentials,
    send_worker_message, add_worker_profile_note,
)

# ── Page config ───────────────────────────────────────────────────────────────

st.set_page_config(page_title="Connect Care", page_icon="🛡️", layout="wide")

st.markdown("""
<style>
    .block-container { padding-top: 1.2rem; padding-bottom: 1rem; }
    .kpi { border-radius:10px; padding:1.1rem 1rem; text-align:center; }
    .kpi-num  { font-size:2.2rem; font-weight:700; line-height:1; }
    .kpi-lbl  { font-size:0.8rem; color:#666; margin-top:0.25rem; }
    .chip { display:inline-block; border-radius:20px; padding:2px 10px;
            font-size:0.78rem; font-weight:600; color:#fff; }
    .row-card { border-left:4px solid #eee; border-radius:6px;
                padding:0.5rem 0.8rem; margin-bottom:0.4rem; background:#fafafa; }
</style>
""", unsafe_allow_html=True)

# ── Constants ─────────────────────────────────────────────────────────────────

SEV_ORDER  = ["CRITICAL", "HIGH", "MEDIUM", "LOW", "INFO"]
SEV_COLOUR = {"CRITICAL":"#d62728","HIGH":"#ff7f0e","MEDIUM":"#e6c200","LOW":"#2ca02c","INFO":"#aec7e8"}
SEV_TINT   = {"CRITICAL":"#fff5f5","HIGH":"#fff8f0","MEDIUM":"#fffdf0","LOW":"#f0fff0","INFO":"#f0f6ff"}
SEV_EMOJI  = {"CRITICAL":"🔴","HIGH":"🟠","MEDIUM":"🟡","LOW":"🟢","INFO":"⚪"}
EXCLUDED   = {"(team)", "(unassigned)"}

PLAIN = {
    "NO CLOCK-IN":                              "Didn't clock in",
    "LATE CLOCK-IN":                            "Late to start shift",
    "EARLY CLOCK-OUT":                          "Left shift early",
    "MISSING CLOCK-OUT":                        "Forgot to clock out",
    "AUTO CLOCK-OUT":                           "System closed shift",
    "OPEN SHIFT":                               "No worker assigned",
    "UNSCHEDULED SHIFT":                        "Unrostered shift",
    "SUSPICIOUSLY SHORT SHIFT":                 "Very short shift",
    "MULTIPLE CLOCK-INS SAME CLIENT/DAY":       "Multiple clock-ins same client",
    "UNDERSTAFFED -- RATIO BREACH":             "Not enough staff",
    "OVERSTAFFED -- POSSIBLE OVERBILLING":      "Too many staff",
    "GPS MISMATCH":                             "Wrong location at clock-in",
    "NO SHIFT NOTES":                           "No notes written",
    "EMPTY NOTES":                              "Notes left blank",
    "INSUFFICIENT NOTES":                       "Notes too short",
    "LATE NOTE SUBMISSION":                     "Notes written late",
    "MISSING SIGNATURE":                        "No participant signature",
    "DUPLICATE/COPY-PASTE NOTES":               "Copy-pasted notes",
    "POSSIBLE AI-GENERATED NOTE":               "Possible AI-written notes",
    "FAILS NDIS STANDARD":                      "Doesn't meet NDIS standard",
    "NOT PLAIN ENGLISH":                        "Unclear notes",
    "NOTE DOESN'T MAKE SENSE":                  "Incoherent notes",
    "SUBJECTIVE LANGUAGE":                      "Opinions instead of observations",
    "NOT PERSON-CENTRED":                       "Not person-centred",
    "INCIDENT KEYWORD -- VERIFY REPORT FILED":  "Incident — check report filed",
    "RESTRICTIVE PRACTICE MENTIONED":           "Restrictive practice — needs authorisation",
    "MEDICATION MENTIONED -- VERIFY FORM FILED":"Medication — check form filed",
    "INCOMPLETE INCIDENT REPORT":               "Incomplete incident report",
    "LATE INCIDENT REPORTING":                  "Late incident report",
    "MISSING FORM -- KALLAN":                   "Kallan form not submitted",
    "MISSING FORM -- EVAN":                     "Evan form not submitted",
    "MISSING FORM -- MICHAEL":                  "Michael form not submitted",
    "FORM FREQUENCY -- JOSHUA":                 "Joshua weekly form missing",
    "FORM FREQUENCY -- NADA":                   "Nada weekly form missing",
    "FORM FREQUENCY -- JOHN":                   "John weekly form missing",
    "FORM FREQUENCY -- NICOLE":                 "Nicole weekly form missing",
    "APPROVED LEAVE":                           "Approved leave",
    "BREAK COMPLIANCE":                         "No break on long shift",
    "CROSS-WORKER COPY-PASTE NOTES":            "Identical notes across workers",
    "ONBOARDING INCOMPLETE":                    "Onboarding not complete",
    "UNAUTHORISED CLIENT ACCESS":               "Unauthorised client access",
    "EXPIRED CREDENTIAL":                       "Credential expired",
    "CREDENTIAL EXPIRING SOON":                 "Credential expiring soon",
}

REQUIRED_DOCS = [
    "NDIS Worker Screening",
    "Working With Children Check",
    "Police Check",
    "First Aid Certificate",
    "CPR Certificate",
    "Manual Handling Training",
]

def plain(cat): return PLAIN.get(cat, cat.title())

# ── Helpers ───────────────────────────────────────────────────────────────────

def parse_date(s):
    today = datetime.date.today()
    for fmt in ["%Y-%m-%d", "%a %d-%b"]:
        try:
            d = datetime.datetime.strptime(s.strip(), fmt).date()
            if fmt == "%a %d-%b":
                d = d.replace(year=today.year)
                if d > today + datetime.timedelta(days=30):
                    d = d.replace(year=today.year - 1)
            return d
        except Exception:
            pass
    return None

def compliance_score(wdf):
    deduct = {"CRITICAL":15,"HIGH":8,"MEDIUM":3,"LOW":1}
    return max(0, 100 - int(wdf["Severity"].map(deduct).fillna(0).sum()))

def score_colour(s): return "#2ca02c" if s>=80 else "#ff7f0e" if s>=60 else "#d62728"
def score_label(s):  return "Good" if s>=80 else "Needs Attention" if s>=60 else "At Risk"

def colour_row(row):
    t = SEV_TINT.get(row["Severity"], "#fff")
    return [f"background-color:{t}"] * len(row)

def to_excel(dfs):
    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        for sheet, df in dfs.items():
            df.to_excel(w, sheet_name=sheet[:31], index=False)
    return buf.getvalue()

def pay_cycles(today):
    y, m = today.year, today.month
    last = calendar.monthrange(y, m)[1]
    mn   = today.strftime("%b %Y")
    pm = m-1 if m>1 else 12; py = y if m>1 else y-1
    plast = calendar.monthrange(py, pm)[1]
    pmn = datetime.date(py, pm, 1).strftime("%b %Y")
    return [
        (f"Pay Cycle 1 — 1–15 {mn}",     datetime.date(y,m,1),    datetime.date(y,m,15)),
        (f"Pay Cycle 2 — 16–{last} {mn}", datetime.date(y,m,16),   datetime.date(y,m,last)),
        (f"Pay Cycle 1 — 1–15 {pmn}",     datetime.date(py,pm,1),  datetime.date(py,pm,15)),
        (f"Pay Cycle 2 — 16–{plast} {pmn}",datetime.date(py,pm,16),datetime.date(py,pm,plast)),
    ]

def doc_status(v):
    if not v or str(v).strip() in ("","nan","None"):
        return "❌","Missing","#d62728"
    try:
        d = datetime.date.fromisoformat(str(v).strip())
        n = (d - datetime.date.today()).days
        if n < 0:   return "❌", f"Expired {abs(n)}d ago","#d62728"
        if n <= 60: return "⚠️", f"Expires in {n}d","#ff7f0e"
        return "✅", f"Valid · {d.strftime('%d %b %Y')}","#2ca02c"
    except Exception:
        return "❓","Invalid date","#888"

# ── Notification store ────────────────────────────────────────────────────────

_NOTIF_FILE = os.path.join(os.path.dirname(__file__), "notifications_log.json")

def _load_notifs():
    try:
        with open(_NOTIF_FILE, "r", encoding="utf-8") as f: return json.load(f)
    except Exception: return []

def _save_notifs(ns):
    try:
        with open(_NOTIF_FILE, "w", encoding="utf-8") as f: json.dump(ns, f, default=str, indent=2)
    except Exception: pass

def _init_notifs():
    if "notifications" not in st.session_state:
        st.session_state.notifications = _load_notifs()

# ── Password gate ─────────────────────────────────────────────────────────────

if not st.session_state.get("authenticated"):
    st.markdown("<br><br>", unsafe_allow_html=True)
    _, col, _ = st.columns([1,1,1])
    with col:
        st.markdown("## 🛡️ Connect Care")
        st.markdown("##### Compliance Dashboard")
        st.markdown("<br>", unsafe_allow_html=True)
        pw = st.text_input("Password", type="password", placeholder="Enter password…")
        if st.button("Sign in", use_container_width=True, type="primary"):
            if pw == st.secrets.get("DASHBOARD_PASSWORD",""):
                st.session_state.authenticated = True
                st.rerun()
            else:
                st.error("Incorrect password.")
    st.stop()

# ── Data loaders ──────────────────────────────────────────────────────────────

_AUDIT_DAYS      = 45
_AUDIT_CACHE_FILE = os.path.join(os.path.dirname(__file__), "audit_cache.json")

def _save_audit_cache(issues, ran_at):
    try:
        data = {
            "ran_at": ran_at,
            "issues": [
                {"severity": i.severity, "category": i.category, "worker": i.worker,
                 "client": i.client or "", "date": i.date or "", "detail": i.detail or ""}
                for i in issues
            ],
        }
        with open(_AUDIT_CACHE_FILE, "w", encoding="utf-8") as f:
            json.dump(data, f)
    except Exception:
        pass

def _load_audit_cache():
    try:
        with open(_AUDIT_CACHE_FILE, "r", encoding="utf-8") as f:
            data = json.load(f)
        from types import SimpleNamespace
        issues = [SimpleNamespace(**d) for d in data["issues"]]
        return issues, data["ran_at"]
    except Exception:
        return None, None

@st.cache_data(ttl=1800, show_spinner=False)
def _fetch_audit():
    return run_audit(_AUDIT_DAYS)

@st.cache_data(ttl=3600, show_spinner=False)
def _load_contacts_raw():
    try:
        users = fetch_all_users()
        return {
            f"{u.get('firstName','')} {u.get('lastName','')}".strip(): {
                "phone":  u.get("phoneNumber") or u.get("phone") or "",
                "userId": u.get("userId"),
            }
            for u in users.values()
        }
    except Exception: return {}

def get_contacts():
    if "contacts" not in st.session_state:
        st.session_state.contacts = _load_contacts_raw()
    return st.session_state.contacts

# Load last audit from disk if this is a fresh session
if "audit_issues" not in st.session_state:
    _cached_issues, _cached_ran_at = _load_audit_cache()
    if _cached_issues is not None:
        st.session_state.audit_issues = _cached_issues
        st.session_state.audit_ran_at = _cached_ran_at

# ── Sidebar ───────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown("### 🛡️ Connect Care")
    st.caption("Compliance Dashboard")
    st.divider()

    today     = datetime.date.today()
    yesterday = today - datetime.timedelta(days=1)
    cycles    = pay_cycles(today)

    period_opts = (
        [f"Yesterday — {yesterday.strftime('%a %d %b')}"]
        + [c[0] for c in cycles]
        + ["Custom range"]
    )
    period_choice = st.selectbox("Period", period_opts, label_visibility="collapsed")

    if period_choice == "Custom range":
        dr = st.date_input("Dates", value=(today - datetime.timedelta(days=7), today),
                           max_value=today, label_visibility="collapsed")
        start_date = dr[0] if isinstance(dr,(list,tuple)) and len(dr)==2 else today-datetime.timedelta(days=7)
        end_date   = dr[1] if isinstance(dr,(list,tuple)) and len(dr)==2 else today
    elif period_choice.startswith("Yesterday"):
        start_date = end_date = yesterday
    else:
        _, start_date, end_date = next(c for c in cycles if c[0]==period_choice)

    st.markdown("<br>", unsafe_allow_html=True)

    if st.button("▶  Run Audit", use_container_width=True, type="primary"):
        _fetch_audit.clear()
        if "contacts" in st.session_state: del st.session_state["contacts"]
        with st.spinner("Running audit… (~20 seconds)"):
            issues = _fetch_audit()
        ran_at = datetime.datetime.now().strftime("%d %b %Y, %I:%M %p")
        st.session_state.audit_issues = issues
        st.session_state.audit_ran_at = ran_at
        _save_audit_cache(issues, ran_at)
        st.rerun()

    ran_at = st.session_state.get("audit_ran_at")
    if ran_at:
        st.markdown(f"<small style='color:#aaa'>Last audit: {ran_at}</small>", unsafe_allow_html=True)
    else:
        st.markdown("<small style='color:#aaa'>No audit run yet</small>", unsafe_allow_html=True)

    st.divider()
    if st.button("Sign out", use_container_width=True):
        st.session_state.authenticated = False
        st.rerun()

all_issues = st.session_state.get("audit_issues")

# ── Build dataframe ───────────────────────────────────────────────────────────

_init_notifs()

if all_issues is None:
    st.markdown("<br><br>", unsafe_allow_html=True)
    _, col, _ = st.columns([1, 2, 1])
    with col:
        st.markdown("### 🛡️ Connect Care Compliance")
        st.markdown("Click **▶ Run Audit** in the sidebar to load shift data from Connecteam.")
        st.markdown("<small style='color:#aaa'>First run takes ~10 seconds. Results stay loaded for 30 minutes.</small>", unsafe_allow_html=True)
    st.stop()

if not all_issues:
    st.success("✅ No compliance issues found for this period.")
    st.stop()

df_raw = pd.DataFrame([
    {"Severity": i.severity, "Category": i.category, "Issue": plain(i.category),
     "Worker": i.worker, "Client": i.client or "", "Date": i.date or "", "Detail": i.detail or ""}
    for i in all_issues
])
df_raw["_d"] = df_raw["Date"].apply(parse_date)
df = df_raw[(df_raw["_d"] >= start_date) & (df_raw["_d"] <= end_date)].drop(columns="_d").copy()
df["_r"] = df["Severity"].map({s:i for i,s in enumerate(SEV_ORDER)})
df = df.sort_values("_r").drop(columns="_r")

df_staff = df[~df["Worker"].isin(EXCLUDED)].copy()

n_crit  = int((df["Severity"]=="CRITICAL").sum())
n_high  = int((df["Severity"]=="HIGH").sum())
n_med   = int((df["Severity"]=="MEDIUM").sum())
n_total = len(df)

period_str = (yesterday.strftime("%a %d %b") if start_date==end_date==yesterday
              else f"{start_date.strftime('%d %b')} – {end_date.strftime('%d %b %Y')}")

# ── Page header ───────────────────────────────────────────────────────────────

st.markdown(f"## Connect Care &nbsp; <small style='font-size:1rem;color:#888;font-weight:400'>{period_str}</small>", unsafe_allow_html=True)

if n_crit:
    st.error(f"⚠️ {n_crit} critical issue{'s' if n_crit>1 else ''} need immediate attention.")

c1,c2,c3,c4 = st.columns(4)
for col, bg, fg, lbl, num in [
    (c1,"#fff5f5","#d62728","🔴 Critical",    n_crit),
    (c2,"#fff8f0","#ff7f0e","🟠 High",        n_high),
    (c3,"#fffdf0","#b8980a","🟡 Medium",      n_med),
    (c4,"#f0f4ff","#4c78a8","📋 Total Issues",n_total),
]:
    col.markdown(
        f'<div class="kpi" style="background:{bg}">'
        f'<div class="kpi-num" style="color:{fg}">{num}</div>'
        f'<div class="kpi-lbl">{lbl}</div></div>',
        unsafe_allow_html=True)

st.markdown("<br>", unsafe_allow_html=True)

# ── Tabs ──────────────────────────────────────────────────────────────────────

tab_overview, tab_workers, tab_clients, tab_docs, tab_messages, tab_export = st.tabs([
    "🚨 Action Required",
    "👤 Workers",
    "👥 Clients",
    "📄 Documents",
    "📬 Messages",
    "📋 Audit Package",
])

# ═══════════════════════════════════════════════════════
# TAB 1 — Action Required
# ═══════════════════════════════════════════════════════
with tab_overview:
    urgent = df[df["Severity"].isin(["CRITICAL","HIGH"])]

    if urgent.empty:
        st.success("✅ No critical or high priority issues this period.")
    else:
        # Group by worker
        for worker in sorted(urgent["Worker"].unique()):
            if worker in EXCLUDED: continue
            wissues = urgent[urgent["Worker"]==worker]
            crit_n = int((wissues["Severity"]=="CRITICAL").sum())
            high_n = int((wissues["Severity"]=="HIGH").sum())
            badge = ""
            if crit_n: badge += f'<span class="chip" style="background:#d62728">{crit_n} Critical</span> '
            if high_n: badge += f'<span class="chip" style="background:#ff7f0e">{high_n} High</span>'
            with st.expander(f"**{worker}** &nbsp; {badge}", expanded=crit_n>0):
                for _, row in wissues.iterrows():
                    c = SEV_COLOUR[row["Severity"]]
                    st.markdown(
                        f'<div class="row-card" style="border-left-color:{c}">'
                        f'{SEV_EMOJI[row["Severity"]]} <strong>{row["Issue"]}</strong>'
                        f' &nbsp;·&nbsp; {row["Client"]} &nbsp;·&nbsp; '
                        f'<span style="color:#888;font-size:0.85rem">{row["Date"]}</span><br>'
                        f'<span style="color:#444;font-size:0.88rem">{row["Detail"]}</span></div>',
                        unsafe_allow_html=True)

    # Special flags
    for cat, title, colour in [
        ("RESTRICTIVE PRACTICE MENTIONED",           "🚫 Restrictive Practice — authorisation required", "#d62728"),
        ("LATE INCIDENT REPORTING",                  "⏰ Late Incident Reports — NDIS risk", "#ff7f0e"),
        ("INCIDENT KEYWORD -- VERIFY REPORT FILED",  "📋 Incident Mentioned — confirm report filed", "#ff7f0e"),
        ("ONBOARDING INCOMPLETE",                    "🎓 Onboarding Incomplete", "#d62728"),
        ("UNAUTHORISED CLIENT ACCESS",               "🔑 Unauthorised Client Access", "#d62728"),
    ]:
        rows = df[df["Category"]==cat]
        if not rows.empty:
            st.markdown(f"<hr style='margin:1rem 0'>", unsafe_allow_html=True)
            st.markdown(f"**{title}** ({len(rows)})")
            for _, r in rows.iterrows():
                st.markdown(f"- **{r['Worker']}** · {r['Client']} · {r['Detail']}")

# ═══════════════════════════════════════════════════════
# TAB 2 — Workers
# ═══════════════════════════════════════════════════════
with tab_workers:
    if df_staff.empty:
        st.info("No staff issues this period.")
    else:
        # Scores table
        ws = df_staff.groupby(["Worker","Severity"]).size().unstack(fill_value=0)
        for col in SEV_ORDER:
            if col not in ws.columns: ws[col] = 0
        ws = ws[[c for c in SEV_ORDER if c in ws.columns]]
        ws["Total"]  = ws.sum(axis=1)
        ws["Score"]  = [compliance_score(df_staff[df_staff["Worker"]==w]) for w in ws.index]
        ws["Status"] = ws["Score"].apply(score_label)
        ws = ws.sort_values("Score")

        st.dataframe(ws, use_container_width=True, column_config={
            "CRITICAL": st.column_config.NumberColumn("🔴"),
            "HIGH":     st.column_config.NumberColumn("🟠"),
            "MEDIUM":   st.column_config.NumberColumn("🟡"),
            "LOW":      st.column_config.NumberColumn("🟢"),
            "Score":    st.column_config.ProgressColumn("Score", min_value=0, max_value=100, format="%d%%"),
        })

        st.divider()

        # Worker drill-down
        selected = st.selectbox("Select a worker to review", ["—"] + sorted(ws.index.tolist()))
        if selected != "—":
            score = compliance_score(df_staff[df_staff["Worker"]==selected])
            sc = score_colour(score)

            col_s, col_info = st.columns([1,3])
            with col_s:
                st.markdown(
                    f'<div style="text-align:center;padding:1.5rem 0">'
                    f'<div style="font-size:2.8rem;font-weight:700;color:{sc}">{score}</div>'
                    f'<div style="color:{sc};font-weight:600">{score_label(score)}</div></div>',
                    unsafe_allow_html=True)
            with col_info:
                contacts = get_contacts()
                info = contacts.get(selected, {})
                phone = info.get("phone","")
                st.markdown("<br>", unsafe_allow_html=True)
                st.markdown(f"📞 {'['+phone+'](tel:'+phone+')' if phone else '_no phone_'}")

            wdf = df_staff[df_staff["Worker"]==selected][["Severity","Issue","Client","Date","Detail"]]
            st.dataframe(wdf.style.apply(colour_row, axis=1),
                         use_container_width=True, hide_index=True)

            # Download
            st.download_button(
                "⬇️ Download issues (CSV)",
                data=wdf.to_csv(index=False),
                file_name=f"{selected.replace(' ','_')}_{period_str.replace(' ','_')}.csv",
                mime="text/csv",
            )

            # Notify
            st.markdown("#### Send message")
            worker_id = info.get("userId")
            if not worker_id:
                st.warning("No Connecteam user ID found for this worker.")
            else:
                default_msg = (
                    f"Hi {selected.split()[0]},\n\n"
                    f"Following up on your shift {period_str.lower()} — there are a few things that need sorting:\n\n"
                    + "\n".join(
                        f"{'🔴' if r['Severity']=='CRITICAL' else '🟠' if r['Severity']=='HIGH' else '🟡'} "
                        f"{r['Client']} — {r['Detail']}"
                        for _,r in wdf.iterrows()
                        if r["Severity"] in ("CRITICAL","HIGH","MEDIUM")
                    )[:6*120]  # cap length
                    + "\n\nCan you let me know what happened?\n\nCheers"
                )
                msg = st.text_area("Message", value=default_msg, height=220, key=f"msg_{selected}")
                if st.button("📤 Send via Connecteam", type="primary", key=f"send_{selected}"):
                    if not CT_CHAT_READY:
                        st.error("CONNECTEAM_SENDER_ID not set in secrets — can't send.")
                    else:
                        ok, err = send_worker_message(worker_id, msg)
                        if ok:
                            sev_counts = wdf["Severity"].value_counts().to_dict()
                            st.session_state.notifications.insert(0, {
                                "id": str(uuid.uuid4())[:8],
                                "worker": selected, "worker_id": worker_id,
                                "sent_at": datetime.datetime.now().strftime("%d %b %Y, %I:%M %p"),
                                "sent_at_iso": datetime.datetime.now().isoformat(),
                                "period": period_str, "severity_counts": sev_counts,
                                "issue_count": len(wdf),
                                "issues": wdf[["Severity","Issue","Client","Date","Detail"]].to_dict("records"),
                                "message_sent": msg, "status": "Sent",
                                "acknowledged_at": None, "resolved_at": None, "manager_notes": "",
                            })
                            _save_notifs(st.session_state.notifications)
                            st.success(f"✅ Message sent to {selected}. Logged in Messages tab.")
                        else:
                            st.error(f"Failed: {err}")

# ═══════════════════════════════════════════════════════
# TAB 3 — Clients
# ═══════════════════════════════════════════════════════
with tab_clients:
    cs = df.groupby(["Client","Severity"]).size().unstack(fill_value=0)
    for col in SEV_ORDER:
        if col not in cs.columns: cs[col] = 0
    cs = cs[[c for c in SEV_ORDER if c in cs.columns]]
    cs["Total"] = cs.sum(axis=1)
    cs = cs.sort_values("CRITICAL" if "CRITICAL" in cs.columns else "Total", ascending=False)

    st.dataframe(cs, use_container_width=True, column_config={
        "CRITICAL": st.column_config.NumberColumn("🔴 Critical"),
        "HIGH":     st.column_config.NumberColumn("🟠 High"),
        "MEDIUM":   st.column_config.NumberColumn("🟡 Medium"),
        "LOW":      st.column_config.NumberColumn("🟢 Low"),
        "Total":    st.column_config.NumberColumn("Total"),
    })

    st.divider()
    selected_client = st.selectbox("Select a client to review", ["—"] + sorted(df["Client"].unique()))
    if selected_client != "—":
        cdf = df[df["Client"]==selected_client][["Severity","Issue","Worker","Date","Detail"]]
        st.dataframe(cdf.style.apply(colour_row, axis=1), use_container_width=True, hide_index=True)
        st.download_button(
            "⬇️ Download (CSV)",
            data=cdf.to_csv(index=False),
            file_name=f"{selected_client.replace(' ','_')}_{period_str.replace(' ','_')}.csv",
            mime="text/csv",
        )

# ═══════════════════════════════════════════════════════
# TAB 4 — Documents
# ═══════════════════════════════════════════════════════
with tab_docs:
    st.caption("Track NDIS-required document expiry dates for each worker. ✅ Valid · ⚠️ Expiring within 60 days · ❌ Expired / missing")

    # Load saved data from secrets (persistent) or session state (in-session edits)
    if "doc_data" not in st.session_state:
        saved = st.secrets.get("DOCUMENTS_JSON","")
        try:    st.session_state.doc_data = json.loads(saved) if saved else {}
        except: st.session_state.doc_data = {}

    # Load staff names from contacts (already cached)
    contacts_for_docs = get_contacts()
    staff_names = sorted(k for k in contacts_for_docs if k and k not in EXCLUDED)

    if not staff_names:
        st.warning("Could not load staff list. Check API key.")
    else:
        rows = [{"Worker": w, **{d: st.session_state.doc_data.get(w,{}).get(d,"") for d in REQUIRED_DOCS}}
                for w in staff_names]
        doc_df = pd.DataFrame(rows)

        edited = st.data_editor(
            doc_df, use_container_width=True, hide_index=True,
            column_config={
                "Worker": st.column_config.TextColumn("Worker", disabled=True, width="medium"),
                **{d: st.column_config.TextColumn(d, width="small") for d in REQUIRED_DOCS}
            },
            key="doc_editor",
        )

        col_save, col_dl = st.columns([1,1])
        with col_save:
            if st.button("💾 Save", type="primary", use_container_width=True):
                st.session_state.doc_data = {
                    row["Worker"]: {d: str(row[d]) if row[d] else "" for d in REQUIRED_DOCS}
                    for _, row in edited.iterrows()
                }
                st.success("Saved for this session.")
        with col_dl:
            export_rows = [
                {"Worker": row["Worker"], "Document": d,
                 "Expiry": row[d], "Status": f"{doc_status(row[d])[0]} {doc_status(row[d])[1]}"}
                for _, row in edited.iterrows() for d in REQUIRED_DOCS
            ]
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as w:
                pd.DataFrame(export_rows).to_excel(w, index=False)
            st.download_button(
                "⬇️ Download Excel", data=buf.getvalue(),
                file_name=f"ConnectCare_Documents_{today.strftime('%d-%b-%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

        # Alerts only
        alerts = [
            {"Worker": row["Worker"], "Document": d, "Status": f"{doc_status(row[d])[0]} {doc_status(row[d])[1]}"}
            for _, row in edited.iterrows() for d in REQUIRED_DOCS
            if doc_status(row[d])[0] in ("❌","⚠️")
        ]
        if alerts:
            st.markdown("---")
            st.markdown(f"**{len(alerts)} document(s) need attention**")
            st.dataframe(pd.DataFrame(alerts), use_container_width=True, hide_index=True)
        else:
            st.success("✅ All documents on file and valid.")

        with st.expander("How to save document data permanently"):
            st.markdown("""
1. Click **Download Excel** and keep a local copy
2. To persist across sessions: go to [share.streamlit.io](https://share.streamlit.io) → your app → ⚙️ Settings → Secrets → add:
```
DOCUMENTS_JSON = '<paste the JSON from "Download JSON" below>'
```
""")
            st.download_button(
                "⬇️ Download JSON backup",
                data=json.dumps(st.session_state.doc_data, indent=2),
                file_name="documents_backup.json",
                mime="application/json",
            )

# ═══════════════════════════════════════════════════════
# TAB 5 — Messages
# ═══════════════════════════════════════════════════════
with tab_messages:
    notifs = st.session_state.get("notifications", [])

    # Summary row
    n_sent = sum(1 for n in notifs if n["status"]=="Sent")
    n_ack  = sum(1 for n in notifs if n["status"]=="Acknowledged")
    n_res  = sum(1 for n in notifs if n["status"]=="Resolved")
    m1,m2,m3,m4 = st.columns(4)
    m1.metric("Total Sent",    len(notifs))
    m2.metric("Awaiting Reply", n_sent)
    m3.metric("Acknowledged",  n_ack)
    m4.metric("Resolved",      n_res)

    st.divider()

    if not notifs:
        st.info("No messages sent yet. Use the Workers tab to notify a worker.")
    else:
        STATUS_COLOUR = {"Sent":"#ff7f0e","Acknowledged":"#4c78a8","Resolved":"#2ca02c"}

        f_status = st.radio("Show", ["All","Awaiting Reply","Acknowledged","Resolved"], horizontal=True)
        fmap = {"All": None, "Awaiting Reply": "Sent", "Acknowledged": "Acknowledged", "Resolved": "Resolved"}
        filtered = notifs if not fmap[f_status] else [n for n in notifs if n["status"]==fmap[f_status]]

        for n in filtered:
            sc = STATUS_COLOUR.get(n["status"],"#888")
            crit = n["severity_counts"].get("CRITICAL",0)
            high = n["severity_counts"].get("HIGH",0)
            with st.expander(
                f"**{n['worker']}** · {n['sent_at']} · "
                f"{'🔴'+str(crit)+' · ' if crit else ''}"
                f"{'🟠'+str(high)+' · ' if high else ''}"
                f"{n['status']}",
                expanded=False
            ):
                col_l, col_r = st.columns([1,1])
                with col_l:
                    st.markdown(f"**Period:** {n['period']}  ·  **Issues:** {n['issue_count']}")
                    for iss in n["issues"][:6]:
                        st.markdown(f"- {SEV_EMOJI.get(iss['Severity'],'•')} {iss['Issue']} · {iss['Client']}")
                    if len(n["issues"])>6: st.caption(f"…+{len(n['issues'])-6} more")
                with col_r:
                    st.code(n["message_sent"], language=None)

                real_idx = notifs.index(n)
                a1,a2 = st.columns(2)
                with a1:
                    if n["status"]=="Sent" and st.button("✅ Mark Acknowledged", key=f"ack_{n['id']}"):
                        notifs[real_idx]["status"] = "Acknowledged"
                        notifs[real_idx]["acknowledged_at"] = datetime.datetime.now().strftime("%d %b %Y, %I:%M %p")
                        _save_notifs(notifs); st.rerun()
                with a2:
                    if n["status"]!="Resolved" and st.button("🏁 Mark Resolved", key=f"res_{n['id']}"):
                        notifs[real_idx]["status"] = "Resolved"
                        notifs[real_idx]["resolved_at"] = datetime.datetime.now().strftime("%d %b %Y, %I:%M %p")
                        if not notifs[real_idx]["acknowledged_at"]:
                            notifs[real_idx]["acknowledged_at"] = notifs[real_idx]["resolved_at"]
                        _save_notifs(notifs); st.rerun()

                note = st.text_input("Notes", value=n.get("manager_notes",""),
                                     placeholder="e.g. Worker called and confirmed…",
                                     key=f"note_{n['id']}")
                if note != n.get("manager_notes",""):
                    notifs[real_idx]["manager_notes"] = note

        st.divider()
        if notifs:
            notif_df = pd.DataFrame([{
                "Worker":       n["worker"],    "Sent":    n["sent_at"],
                "Period":       n["period"],    "Issues":  n["issue_count"],
                "Status":       n["status"],    "Notes":   n.get("manager_notes",""),
            } for n in notifs])
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as w:
                notif_df.to_excel(w, index=False)
            st.download_button(
                "⬇️ Download message log",
                data=buf.getvalue(),
                file_name=f"ConnectCare_Messages_{today.strftime('%d-%b-%Y')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )

# ═══════════════════════════════════════════════════════
# TAB 6 — Audit Package Export
# ═══════════════════════════════════════════════════════
with tab_export:
    st.markdown("### 📋 NDIS Audit Evidence Package")
    st.caption(
        "Generate a formatted Excel workbook with all compliance evidence for an NDIS audit "
        "or internal review. Select the date range below, then click Download."
    )

    if not _OPENPYXL_OK:
        st.error("openpyxl is not installed. Run `pip install openpyxl` to enable audit package export.")
        st.stop()

    # ── Date range selector ───────────────────────────────────────────────────
    st.markdown("#### Date range")
    export_col1, export_col2 = st.columns(2)
    with export_col1:
        export_start = st.date_input(
            "From",
            value=today - datetime.timedelta(days=30),
            max_value=today,
            key="export_start",
        )
    with export_col2:
        export_end = st.date_input(
            "To",
            value=today,
            min_value=export_start,
            max_value=today,
            key="export_end",
        )

    export_period = f"{export_start.strftime('%d %b %Y')} – {export_end.strftime('%d %b %Y')}"
    st.caption(f"Selected period: **{export_period}**")
    st.markdown("<br>", unsafe_allow_html=True)

    # ── Excel builder ─────────────────────────────────────────────────────────

    def _date_in_range(date_str: str) -> bool:
        """Return True if date_str (various formats) falls within [export_start, export_end]."""
        d = parse_date(date_str)
        if d is None:
            return True   # unknown date — include rather than silently drop
        return export_start <= d <= export_end

    SEV_FILL = {
        "CRITICAL": PatternFill("solid", fgColor="FFD0D0"),
        "HIGH":     PatternFill("solid", fgColor="FFE5CC"),
        "MEDIUM":   PatternFill("solid", fgColor="FFFACC"),
    }
    HEADER_FILL = PatternFill("solid", fgColor="D9E8FB")
    HEADER_FONT = Font(bold=True, color="1A1A2E")

    def _style_sheet(ws, header_row: int = 1):
        """Apply header style, freeze pane, and auto-width to an openpyxl worksheet."""
        for cell in ws[header_row]:
            cell.font      = HEADER_FONT
            cell.fill      = HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        for col_idx, col_cells in enumerate(ws.columns, 1):
            max_len = max(
                (len(str(c.value or "")) for c in col_cells),
                default=8,
            )
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max(max_len + 2, 12), 55)

    def _apply_severity_fills(ws, sev_col_idx: int, header_row: int = 1):
        """Colour-fill rows based on severity value in the given column (1-indexed)."""
        for row in ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row):
            sev_cell = row[sev_col_idx - 1]
            fill     = SEV_FILL.get(str(sev_cell.value or "").upper())
            if fill:
                for cell in row:
                    cell.fill = fill

    def build_audit_package() -> bytes:
        """Assemble all sheets and return the workbook as bytes."""
        import re as _re
        import hashlib as _hashlib

        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # remove default empty sheet

        # ── Filter audit issues to selected date range ────────────────────────
        period_issues = [
            i for i in (all_issues or [])
            if _date_in_range(i.date or "")
        ]

        # ── Sheet 1: Shift Attendance ─────────────────────────────────────────
        ws1 = wb.create_sheet("Shift Attendance")
        ATTEND_HEADERS = [
            "Date", "Worker", "Client",
            "Scheduled Start", "Actual Clock-In",
            "Scheduled End",   "Actual Clock-Out",
            "Clock-In GPS Status", "Duration Hours", "Notes Submitted (Y/N)",
        ]
        ws1.append(ATTEND_HEADERS)

        ATTEND_CATS = {
            "NO CLOCK-IN", "LATE CLOCK-IN", "EARLY CLOCK-OUT",
            "MISSING CLOCK-OUT", "AUTO CLOCK-OUT", "GPS MISMATCH",
            "GPS DATA MISSING", "SUSPICIOUSLY SHORT SHIFT",
            "UNSCHEDULED SHIFT", "MULTIPLE CLOCK-INS SAME CLIENT/DAY",
        }
        NOTE_CATS = {"NO SHIFT NOTES", "EMPTY NOTES", "INSUFFICIENT NOTES"}

        # Build a set of (worker, date) combos with note issues
        no_notes_set = {
            (i.worker, i.date or "") for i in period_issues if i.category in NOTE_CATS
        }

        seen_attend = set()
        for iss in period_issues:
            if iss.category not in ATTEND_CATS:
                continue
            key = (iss.worker, iss.date or "", iss.category)
            if key in seen_attend:
                continue
            seen_attend.add(key)

            # Parse GPS status from detail
            gps_ok = "OK"
            if "GPS" in iss.category:
                gps_ok = "FAIL — " + iss.detail[:60]

            has_notes = "N" if (iss.worker, iss.date or "") in no_notes_set else "Y"

            # Parse times from detail if available (format: "HH:MM–HH:MM" or "arrived HH:MM")
            sched_start = sched_end = actual_in = actual_out = ""
            dur_hours   = ""
            detail = iss.detail or ""
            time_pairs = _re.findall(r"\b(\d{1,2}:\d{2})\b", detail)
            if len(time_pairs) >= 2:
                sched_start, actual_in = time_pairs[0], time_pairs[1]
            elif len(time_pairs) == 1:
                sched_start = time_pairs[0]

            ws1.append([
                iss.date or "", iss.worker, iss.client or "",
                sched_start, actual_in,
                sched_end,   actual_out,
                gps_ok, dur_hours, has_notes,
            ])

        _style_sheet(ws1)

        # ── Sheet 2: Notes Quality ────────────────────────────────────────────
        ws2 = wb.create_sheet("Notes Quality")
        NOTES_HEADERS = [
            "Date", "Worker", "Client",
            "Note Length (words)", "Passes NDIS Standard (Y/N)",
            "Issues Found", "Severity",
        ]
        ws2.append(NOTES_HEADERS)

        NOTES_CATS = {
            "NO SHIFT NOTES":          ("", "N", "No notes at all"),
            "EMPTY NOTES":             ("", "N", "Notes field blank"),
            "INSUFFICIENT NOTES":      ("", "N", "Too short"),
            "FAILS NDIS STANDARD":     ("", "N", ""),
            "POSSIBLE AI-GENERATED NOTE": ("", "?", "Possible AI/template content"),
            "NOT PLAIN ENGLISH":       ("", "?", "Unclear language"),
            "NOTE DOESN'T MAKE SENSE": ("", "?", "Incoherent"),
            "SUBJECTIVE LANGUAGE":     ("", "Y", "Opinions used instead of observations"),
            "NOT PERSON-CENTRED":      ("", "Y", "Not person-centred"),
            "LATE NOTE SUBMISSION":    ("", "Y", "Submitted late — possible backdating"),
            "DUPLICATE/COPY-PASTE NOTES": ("", "?", "Copy-pasted across shifts"),
            "CROSS-WORKER COPY-PASTE NOTES": ("", "?", "Identical notes across two workers"),
        }
        for iss in period_issues:
            if iss.category not in NOTES_CATS:
                continue
            _, ndis_pass, default_issue = NOTES_CATS[iss.category]
            issues_text = iss.detail or default_issue
            # Extract word count if present in detail
            wc_match = _re.search(r"Only (\d+) words", iss.detail or "")
            wc = wc_match.group(1) if wc_match else ""
            row = [
                iss.date or "", iss.worker, iss.client or "",
                wc, ndis_pass or "N",
                issues_text[:200], iss.severity,
            ]
            ws2.append(row)

        _style_sheet(ws2)
        _apply_severity_fills(ws2, sev_col_idx=7)

        # ── Sheet 3: Credentials ─────────────────────────────────────────────
        ws3 = wb.create_sheet("Credentials")
        CRED_HEADERS = [
            "Worker Name", "Credential Type", "Expiry Date",
            "Status", "Days Until Expiry",
        ]
        ws3.append(CRED_HEADERS)

        # Pull credential data fresh (already cached by lru_cache in this session)
        try:
            users_for_creds = fetch_all_users()
            cred_map        = fetch_worker_credentials(users_for_creds)
        except Exception:
            cred_map        = {}
            users_for_creds = {}

        today_d = datetime.date.today()

        def _uname_cred(uid):
            u = users_for_creds.get(uid)
            return f"{u['firstName']} {u['lastName']}" if u else f"User {uid}"

        for uid, creds in sorted(cred_map.items(), key=lambda kv: _uname_cred(kv[0])):
            wname = _uname_cred(uid)
            for cred_type, expiry_date in sorted(creds.items()):
                days_left = (expiry_date - today_d).days
                if days_left < 0:
                    status = "Expired"
                elif days_left <= 30:
                    status = "Expiring Soon"
                else:
                    status = "Valid"
                ws3.append([
                    wname, cred_type,
                    expiry_date.strftime("%d %b %Y"),
                    status, days_left,
                ])

        _style_sheet(ws3)
        # Colour expired rows red, expiring-soon orange
        STATUS_COL = 4   # "Status" is column 4
        for row in ws3.iter_rows(min_row=2, max_row=ws3.max_row):
            status_val = str(row[STATUS_COL - 1].value or "")
            if status_val == "Expired":
                for cell in row:
                    cell.fill = SEV_FILL["CRITICAL"]
            elif status_val == "Expiring Soon":
                for cell in row:
                    cell.fill = SEV_FILL["HIGH"]

        # ── Sheet 4: Compliance Issues ────────────────────────────────────────
        ws4 = wb.create_sheet("Compliance Issues")
        ISSUES_HEADERS = [
            "Date Detected", "Worker", "Client",
            "Issue Category", "Severity", "Detail",
            "Notified (Y/N)", "Acknowledged (Y/N)",
        ]
        ws4.append(ISSUES_HEADERS)

        # Load notified_issues.json for cross-reference
        _notified_path = os.path.join(os.path.dirname(__file__), "notified_issues.json")
        try:
            with open(_notified_path, "r", encoding="utf-8") as _f:
                _notified_raw = json.load(_f)
        except Exception:
            _notified_raw = {}

        # Build a quick set of fingerprints that have been notified
        def _fp(worker, category, date):
            return _hashlib.md5(f"{worker}|{category}|{date or ''}".encode()).hexdigest()

        notified_fps = set(_notified_raw.keys())

        for iss in period_issues:
            fp       = _fp(iss.worker, iss.category, iss.date)
            notified = "Y" if fp in notified_fps else "N"
            acked_v  = _notified_raw.get(fp, {})
            acked    = "Y" if isinstance(acked_v, dict) and acked_v.get("acknowledged") else "N"
            ws4.append([
                iss.date or "",
                iss.worker,
                iss.client or "",
                iss.category,
                iss.severity,
                (iss.detail or "")[:300],
                notified,
                acked,
            ])

        _style_sheet(ws4)
        _apply_severity_fills(ws4, sev_col_idx=5)

        # ── Sheet 5: Notifications Log ────────────────────────────────────────
        ws5 = wb.create_sheet("Notifications Log")
        NOTIF_HEADERS = [
            "Date Sent", "Worker", "Issue Summary",
            "Channel", "Status", "Acknowledged At",
        ]
        ws5.append(NOTIF_HEADERS)

        # Load the dashboard's own notifications_log.json
        _notif_log_path = os.path.join(os.path.dirname(__file__), "notifications_log.json")
        try:
            with open(_notif_log_path, "r", encoding="utf-8") as _f:
                _notif_log = json.load(_f)
        except Exception:
            _notif_log = []

        for entry in _notif_log:
            sent_at = entry.get("sent_at", "")
            # Check date filter
            try:
                sent_dt = datetime.datetime.strptime(sent_at, "%d %b %Y, %I:%M %p")
                if not (export_start <= sent_dt.date() <= export_end):
                    continue
            except Exception:
                pass   # include if we can't parse the date

            summary = "; ".join(
                f"{iss.get('Issue','?')} ({iss.get('Client','?')})"
                for iss in (entry.get("issues") or [])[:3]
            )
            ws5.append([
                sent_at,
                entry.get("worker", ""),
                summary,
                "Connecteam Chat",
                entry.get("status", ""),
                entry.get("acknowledged_at") or "",
            ])

        _style_sheet(ws5)

        # ── Return bytes ──────────────────────────────────────────────────────
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # ── Generate and offer download ───────────────────────────────────────────
    st.info(
        "The audit package pulls data from the most recent audit run. "
        "Click **▶ Run Audit** in the sidebar first if you need fresh data."
    )

    if all_issues is None:
        st.warning("No audit data loaded yet — run the audit from the sidebar first.")
    else:
        n_in_range = sum(
            1 for i in all_issues if _date_in_range(i.date or "")
        )
        st.markdown(f"**{n_in_range}** compliance issues in selected period will be included.")

        generate_btn = st.button(
            "📦 Generate Audit Package",
            type="primary",
            use_container_width=True,
        )

        if generate_btn:
            with st.spinner("Building Excel workbook (credentials + issues + logs)…"):
                try:
                    excel_bytes = build_audit_package()
                    st.session_state["audit_package_bytes"]    = excel_bytes
                    st.session_state["audit_package_period"]   = export_period
                    st.session_state["audit_package_filename"] = (
                        f"ConnectCare_AuditPackage_"
                        f"{export_start.strftime('%d%b')}-{export_end.strftime('%d%b%Y')}.xlsx"
                    )
                    st.success("Audit package ready — click Download below.")
                except Exception as _e:
                    st.error(f"Failed to build audit package: {_e}")

        if st.session_state.get("audit_package_bytes"):
            st.download_button(
                label="📥 Download NDIS Audit Package",
                data=st.session_state["audit_package_bytes"],
                file_name=st.session_state.get(
                    "audit_package_filename",
                    f"ConnectCare_AuditPackage_{today.strftime('%d-%b-%Y')}.xlsx",
                ),
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
            pkg_period = st.session_state.get("audit_package_period", "")
            if pkg_period:
                st.caption(f"Package covers: {pkg_period}")

        st.divider()
        st.markdown("#### What's in the package?")
        st.markdown("""
| Sheet | Contents |
|---|---|
| **Shift Attendance** | Clock-in/out times, GPS status, notes completion |
| **Notes Quality** | Word counts, NDIS standard compliance, AI-detection flags |
| **Credentials** | All worker credential expiry dates with status |
| **Compliance Issues** | Every flagged issue with severity, notification status, and acknowledgement |
| **Notifications Log** | Record of all compliance messages sent to workers |

Rows are colour-coded: 🔴 red = Critical, 🟠 orange = High, 🟡 yellow = Medium.
Headers are bold with a blue tint. Top row is frozen for easy scrolling.
        """)
